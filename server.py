from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone
import base64
import io
import json
import re
import asyncio
from PIL import Image
import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')


# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Models
class RowData(BaseModel):
    row_number: int
    unit_weight: Optional[float] = None
    quantity: Optional[int] = None
    quantity_raw: Optional[str] = None
    row_total: Optional[float] = None
    confidence: str = "high"  # high, medium, low
    has_warning: bool = False
    warning_message: Optional[str] = None

class PageData(BaseModel):
    page_number: int
    rows: List[RowData] = []
    page_total: float = 0.0
    reported_total: Optional[float] = None
    total_difference: Optional[float] = None
    confidence: str = "high"
    thumbnail_base64: Optional[str] = None

class ProcessingResult(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    filename: str
    total_pages: int
    pages: List[PageData]
    grand_total: float
    reported_grand_total: Optional[float] = None
    grand_total_difference: Optional[float] = None
    processing_time: float
    timestamp: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class RowUpdate(BaseModel):
    page_number: int
    row_number: int
    unit_weight: Optional[float] = None
    quantity: Optional[int] = None

class BulkRowUpdate(BaseModel):
    updates: List[RowUpdate]

# Helper functions
def parse_quantity(quantity_str: str) -> int:
    # Extract the FIRST number from quantity format like '2/2', '17/1', '19/19'
    if not quantity_str:
        return 0
    # Clean the string
    quantity_str = quantity_str.strip()
    # Match pattern like "2/2", "17/1" - extract first number
    match = re.match(r'(\d+)\s*/\s*\d+', quantity_str)
    if match:
        return int(match.group(1))
    # Try to extract any number
    numbers = re.findall(r'\d+', quantity_str)
    if numbers:
        return int(numbers[0])
    return 0

def normalize_weight(weight_str: str) -> float:
    # Normalize weight string to float, handling OCR errors
    if not weight_str:
        return 0.0
    # Convert to lowercase and clean
    weight_str = weight_str.lower().strip()
    # Handle common OCR errors: Kq, Ko -> kg
    weight_str = re.sub(r'k[qo]', 'kg', weight_str)
    # Remove 'kg' suffix
    weight_str = re.sub(r'\s*kg\s*$', '', weight_str, flags=re.IGNORECASE)
    # Replace comma with dot for decimal
    weight_str = weight_str.replace(',', '.')
    # Extract numeric value
    match = re.search(r'[\d.]+', weight_str)
    if match:
        try:
            return float(match.group())
        except ValueError:
            return 0.0
    return 0.0

def image_to_base64(image_bytes: bytes) -> str:
    # Convert image bytes to base64 string
    return base64.b64encode(image_bytes).decode('utf-8')

def create_thumbnail(image_bytes: bytes, max_size: int = 200) -> str:
    # Create a thumbnail from image bytes and return as base64
    try:
        img = Image.open(io.BytesIO(image_bytes))
        img.thumbnail((max_size, max_size), Image.Resampling.LANCZOS)
        # Convert to RGB if necessary
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')
        buffer = io.BytesIO()
        img.save(buffer, format='JPEG', quality=60)
        return base64.b64encode(buffer.getvalue()).decode('utf-8')
    except Exception as e:
        logger.error(f"Error creating thumbnail: {e}")
        return ""

async def extract_pdf_pages_as_images(pdf_bytes: bytes) -> List[bytes]:
    # Extract pages from PDF as images
    images = []
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            # Higher resolution for better OCR
            mat = fitz.Matrix(2, 2)  # 2x zoom for better quality
            pix = page.get_pixmap(matrix=mat)
            img_bytes = pix.tobytes("png")
            images.append(img_bytes)
        doc.close()
    except Exception as e:
        logger.error(f"Error extracting PDF pages: {e}")
    return images

async def analyze_image_with_gpt_vision(image_base64: str):
    return {
        "rows": [],
        "reported_total": None,
        "confidence": "low",
        "notes": "AI temporarily disabled"
    }
    

IMPORTANT RULES:
1. Focus ONLY on the main data table
2. IGNORE: titles, headers, QR codes, drawings, margins
3. For each row, extract:
   - Unit weight (in kg, format like "0.20 Kg", "1.63 Kg")
   - Quantity (format like "2/2", "17/1", "19/19")
4. If you find a "Total" or "Total weight" value at the bottom, include it as "reported_total"
5. Handle OCR variations: Kg, Kq, Ko all mean kg

Return ONLY valid JSON in this exact format:
{
  "rows": [
    {"unit_weight": "0.20 Kg", "quantity": "2/2"},
    {"unit_weight": "1.63 Kg", "quantity": "17/1"}
  ],
  "reported_total": "35.47 kg",
  "confidence": "high",
  "notes": "any extraction notes"
}

If no table is found or data is unreadable, return:
{
  "rows": [],
  "reported_total": null,
  "confidence": "low",
  "notes": "reason for failure"
}
    ).with_model("openai", "gpt-5.2")
    
    try:
        image_content = ImageContent(image_base64=image_base64)
        user_message = UserMessage(
            text="Extract the cutting report table data from this image. Return ONLY valid JSON.",
            image_contents=[image_content]
        )
        
        response = await chat.send_message(user_message)
        
        # Parse the JSON response
        # Try to extract JSON from the response
        json_match = re.search(r'\{[\s\S]*\}', response)
        if json_match:
            result = json.loads(json_match.group())
            return result
        else:
            return {
                "rows": [],
                "reported_total": None,
                "confidence": "low",
                "notes": "Could not parse response as JSON"
            }
    except json.JSONDecodeError as e:
        logger.error(f"JSON parse error: {e}")
        return {
            "rows": [],
            "reported_total": None,
            "confidence": "low",
            "notes": f"JSON parse error: {str(e)}"
        }
    except Exception as e:
        logger.error(f"GPT Vision error: {e}")
        return {
            "rows": [],
            "reported_total": None,
            "confidence": "low",
            "notes": f"Error: {str(e)}"
        }

def process_extracted_data(extracted: Dict[str, Any], page_number: int) -> PageData:
    # Process extracted data into PageData format
    rows = []
    page_total = 0.0
    overall_confidence = extracted.get("confidence", "medium")
    
    for idx, row in enumerate(extracted.get("rows", [])):
        unit_weight_str = row.get("unit_weight", "")
        quantity_str = row.get("quantity", "")
        
        unit_weight = normalize_weight(unit_weight_str)
        quantity = parse_quantity(quantity_str)
        row_total = unit_weight * quantity if unit_weight and quantity else 0.0
        
        # Determine row confidence and warnings
        row_confidence = "high"
        has_warning = False
        warning_message = None
        
        if not unit_weight or not quantity:
            row_confidence = "low"
            has_warning = True
            warning_message = "Missing weight or quantity"
        elif unit_weight > 100:  # Suspicious: very high weight
            row_confidence = "medium"
            has_warning = True
            warning_message = "Unusually high weight value"
        elif unit_weight < 0.01:  # Suspicious: very low weight
            row_confidence = "medium"
            has_warning = True
            warning_message = "Unusually low weight value"
        
        row_data = RowData(
            row_number=idx + 1,
            unit_weight=unit_weight,
            quantity=quantity,
            quantity_raw=quantity_str,
            row_total=round(row_total, 3),
            confidence=row_confidence,
            has_warning=has_warning,
            warning_message=warning_message
        )
        rows.append(row_data)
        page_total += row_total
    
    # Process reported total
    reported_total = None
    total_difference = None
    if extracted.get("reported_total"):
        reported_total = normalize_weight(str(extracted.get("reported_total")))
        if reported_total > 0:
            total_difference = round(abs(page_total - reported_total), 3)
            if total_difference > 0.5:  # More than 0.5 kg difference
                overall_confidence = "low" if total_difference > 2 else "medium"
    
    return PageData(
        page_number=page_number,
        rows=rows,
        page_total=round(page_total, 3),
        reported_total=reported_total,
        total_difference=total_difference,
        confidence=overall_confidence
    )

async def process_single_page(image_bytes: bytes, page_number: int) -> PageData:
    # Process a single page image
    image_base64 = image_to_base64(image_bytes)
    thumbnail = create_thumbnail(image_bytes)
    
    extracted = await analyze_image_with_gpt_vision(image_base64)
    page_data = process_extracted_data(extracted, page_number)
    page_data.thumbnail_base64 = thumbnail
    
    return page_data

# API Routes
@api_router.get("/")
async def root():
    return {"message": "Industrial Cutting Report Processor API"}

@api_router.get("/health")
async def health():
    return {"status": "healthy", "timestamp": datetime.now(timezone.utc).isoformat()}

@api_router.post("/process", response_model=ProcessingResult)
async def process_document(file: UploadFile = File(...)):
    # Process uploaded PDF or image file
    import time
    start_time = time.time()
    
    filename = file.filename or "unknown"
    content_type = file.content_type or ""
    file_bytes = await file.read()
    
    logger.info(f"Processing file: {filename}, type: {content_type}, size: {len(file_bytes)} bytes")
    
    pages_data = []
    
    try:
        # Determine if PDF or image
        if content_type == "application/pdf" or filename.lower().endswith('.pdf'):
            # Extract pages from PDF
            page_images = await extract_pdf_pages_as_images(file_bytes)
            logger.info(f"Extracted {len(page_images)} pages from PDF")
            
            # Process pages in parallel (up to 5 at a time)
            tasks = []
            for idx, img_bytes in enumerate(page_images):
                tasks.append(process_single_page(img_bytes, idx + 1))
            
            # Process in batches for better performance
            batch_size = 5
            for i in range(0, len(tasks), batch_size):
                batch = tasks[i:i+batch_size]
                results = await asyncio.gather(*batch)
                pages_data.extend(results)
        else:
            # Single image
            page_data = await process_single_page(file_bytes, 1)
            pages_data.append(page_data)
        
        # Calculate grand total
        grand_total = sum(page.page_total for page in pages_data)
        
        # Check for reported grand total
        reported_grand_total = None
        grand_total_difference = None
        reported_totals = [p.reported_total for p in pages_data if p.reported_total]
        if reported_totals:
            reported_grand_total = sum(reported_totals)
            grand_total_difference = round(abs(grand_total - reported_grand_total), 3)
        
        processing_time = round(time.time() - start_time, 2)
        
        result = ProcessingResult(
            filename=filename,
            total_pages=len(pages_data),
            pages=pages_data,
            grand_total=round(grand_total, 3),
            reported_grand_total=reported_grand_total,
            grand_total_difference=grand_total_difference,
            processing_time=processing_time
        )
        
        # Store in database
        doc = result.model_dump()
        await db.processing_results.insert_one(doc)
        
        return result
        
    except Exception as e:
        logger.error(f"Processing error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/process-multiple")
async def process_multiple_images(files: List[UploadFile] = File(...)):
    # Process multiple image files as separate pages
    import time
    start_time = time.time()
    
    pages_data = []
    
    try:
        # Process each file as a page
        tasks = []
        for idx, file in enumerate(files):
            file_bytes = await file.read()
            tasks.append(process_single_page(file_bytes, idx + 1))
        
        # Process in parallel
        batch_size = 5
        for i in range(0, len(tasks), batch_size):
            batch = tasks[i:i+batch_size]
            results = await asyncio.gather(*batch)
            pages_data.extend(results)
        
        # Calculate grand total
        grand_total = sum(page.page_total for page in pages_data)
        
        # Check for reported grand total
        reported_grand_total = None
        grand_total_difference = None
        reported_totals = [p.reported_total for p in pages_data if p.reported_total]
        if reported_totals:
            reported_grand_total = sum(reported_totals)
            grand_total_difference = round(abs(grand_total - reported_grand_total), 3)
        
        processing_time = round(time.time() - start_time, 2)
        
        filename = f"{len(files)}_images_combined"
        
        result = ProcessingResult(
            filename=filename,
            total_pages=len(pages_data),
            pages=pages_data,
            grand_total=round(grand_total, 3),
            reported_grand_total=reported_grand_total,
            grand_total_difference=grand_total_difference,
            processing_time=processing_time
        )
        
        # Store in database
        doc = result.model_dump()
        await db.processing_results.insert_one(doc)
        
        return result
        
    except Exception as e:
        logger.error(f"Processing error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/update-row")
async def update_row(update: RowUpdate):
    # Update a single row and recalculate totals
    # This endpoint allows manual correction
    return {
        "page_number": update.page_number,
        "row_number": update.row_number,
        "unit_weight": update.unit_weight,
        "quantity": update.quantity,
        "row_total": round((update.unit_weight or 0) * (update.quantity or 0), 3)
    }

@api_router.post("/export-excel")
async def export_excel(data: Dict[str, Any]):
    # Export processing result to Excel
    try:
        wb = openpyxl.Workbook()
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0033CC", end_color="0033CC", fill_type="solid")
        total_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
        warning_fill = PatternFill(start_color="FEF3CD", end_color="FEF3CD", fill_type="solid")
        grand_total_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        grand_total_font = Font(bold=True, color="FFFFFF", size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Summary sheet
        summary_ws = wb.active
        summary_ws.title = "Summary"
        
        summary_ws['A1'] = "Industrial Cutting Report Summary"
        summary_ws['A1'].font = Font(bold=True, size=16)
        summary_ws.merge_cells('A1:D1')
        
        summary_ws['A3'] = "Filename:"
        summary_ws['B3'] = data.get('filename', 'Unknown')
        summary_ws['A4'] = "Total Pages:"
        summary_ws['B4'] = data.get('total_pages', 0)
        summary_ws['A5'] = "Processing Time:"
        summary_ws['B5'] = f"{data.get('processing_time', 0)} seconds"
        
        summary_ws['A7'] = "GRAND TOTAL:"
        summary_ws['B7'] = f"{data.get('grand_total', 0)} kg"
        summary_ws['A7'].font = grand_total_font
        summary_ws['B7'].font = grand_total_font
        summary_ws['A7'].fill = grand_total_fill
        summary_ws['B7'].fill = grand_total_fill
        
        if data.get('reported_grand_total'):
            summary_ws['A8'] = "Reported Total:"
            summary_ws['B8'] = f"{data.get('reported_grand_total')} kg"
            summary_ws['A9'] = "Difference:"
            summary_ws['B9'] = f"{data.get('grand_total_difference', 0)} kg"
        
        # Page sheets
        for page in data.get('pages', []):
            page_num = page.get('page_number', 1)
            ws = wb.create_sheet(title=f"Page {page_num}")
            
            # Headers
            headers = ['Row #', 'Unit Weight (kg)', 'Quantity', 'Quantity (Raw)', 'Row Total (kg)', 'Confidence', 'Warning']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Data rows
            for row_idx, row in enumerate(page.get('rows', []), 2):
                ws.cell(row=row_idx, column=1, value=row.get('row_number'))
                ws.cell(row=row_idx, column=2, value=row.get('unit_weight'))
                ws.cell(row=row_idx, column=3, value=row.get('quantity'))
                ws.cell(row=row_idx, column=4, value=row.get('quantity_raw', ''))
                ws.cell(row=row_idx, column=5, value=row.get('row_total'))
                ws.cell(row=row_idx, column=6, value=row.get('confidence', 'high'))
                ws.cell(row=row_idx, column=7, value=row.get('warning_message', ''))
                
                # Apply warning styling
                if row.get('has_warning'):
                    for col in range(1, 8):
                        ws.cell(row=row_idx, column=col).fill = warning_fill
                
                # Apply borders
                for col in range(1, 8):
                    ws.cell(row=row_idx, column=col).border = border
            
            # Page total row
            total_row = len(page.get('rows', [])) + 2
            ws.cell(row=total_row, column=1, value="PAGE TOTAL")
            ws.cell(row=total_row, column=5, value=page.get('page_total', 0))
            for col in range(1, 8):
                ws.cell(row=total_row, column=col).fill = total_fill
                ws.cell(row=total_row, column=col).font = Font(bold=True)
                ws.cell(row=total_row, column=col).border = border
            
            # Reported total if exists
            if page.get('reported_total'):
                rep_row = total_row + 1
                ws.cell(row=rep_row, column=1, value="Reported Total")
                ws.cell(row=rep_row, column=5, value=page.get('reported_total'))
                ws.cell(row=rep_row + 1, column=1, value="Difference")
                ws.cell(row=rep_row + 1, column=5, value=page.get('total_difference', 0))
            
            # Adjust column widths
            for col in range(1, 8):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"cutting_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Excel export error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/history", response_model=List[Dict[str, Any]])
async def get_processing_history():
    
    results = await db.processing_results.find(
        {},
        {"_id": 0, "pages.thumbnail_base64": 0}  # Exclude large data
    ).sort("timestamp", -1).limit(20).to_list(20)
    return results

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
